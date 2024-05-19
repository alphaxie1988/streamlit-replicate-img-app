import streamlit as st
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
import copy
from minio import Minio
import io
import json
import random
from openpyxl.styles import Alignment

#Config
num_permutations = 5000
bucket_name = "schedule"
object_name = "schedule"
loginUser = "Xie Jianlong (11724)"

#INIT
st.set_page_config(layout="wide",page_title="Scheduler",page_icon="📅")
st.write("🤵"+loginUser)
st.header("Update Preference")
#Connect With MINIO
client = Minio(
"play.min.io",
access_key="minioadmin",
secret_key="minioadmin",
)
#Try to read preferences
try:
    print("GETTING")
    content = client.get_object(bucket_name, object_name=object_name).read().decode('utf-8','ignore')
    preferences =  json.loads(content)
except:
    preferences = {
        'Alpha':[],
        'Bravo':[],
        'Charlie':[],
        'Delta':[],
        'Echo':[],
        'Foxtrot':[],
        'Golf':[],
        'Hotel':[],
        'India':[],
        'Juliett':[],
        'Kilo':[],
        'Lima':[],
        'Mike':[],
        'November':[],
        'Oscar':[],
        'Papa':[],
        'Quebec':[],
        'Romeo':[],
        'Sierra':[],
        'Tango':[],
        'Uniform':[],
        'Victor':[],
        'Whiskey':[],
        'X-ray':[],
        'Yankee':[],
        'Zulu':[]}
if not client.bucket_exists(bucket_name):
    # Pre Create Preferences
    client.make_bucket(bucket_name) #TODO: Update add region
    client.put_object(bucket_name, object_name,content_type="application/json", data=io.BytesIO(json.dumps(preferences).encode('utf-8')), part_size=10*1024*1024,length=-1)

## Manage Editable State
if 'editable' not in st.session_state:
    st.session_state['editable'] = "Not Lock"

if os.path.exists("editable"):
    with open("editable", 'r') as file:
        file_contents = file.read()
        st.session_state['editable'] = True
        if loginUser == file_contents:
            st.session_state['editable'] = "Editable"
        else:
            st.session_state['editable'] = "Lock by "+ file_contents

#Define Function
def lockAndEdit():
    if os.path.exists("editable") != True:
        with open("editable", 'w') as file:
            file.write(loginUser)
    
def savePreference(selected_likes):
    try:
        os.remove("editable")
    except:
        pass
    st.session_state['editable'] = "Not Lock"
    st.toast("Preference Saved", icon='😍')
    preferences[selected_person] = selected_likes
    print("PUTTING")
    client.put_object(bucket_name, object_name,content_type="application/json", data=io.BytesIO(json.dumps(preferences).encode('utf-8')), part_size=10*1024*1024,length=-1)

try:
    editableAnot = st.session_state['editable'] == "Editable"
except:
    editableAnot = True

selected_person = st.selectbox('Select a person:', list(preferences.keys()),disabled=editableAnot)
if selected_person in preferences:
    selected_preferences = preferences[selected_person]
    available_options = list(preferences.keys())
    filtered_options = [option for option in available_options if option != selected_person]
    selected_likes = st.multiselect('Select people that ' + selected_person + ' likes:', filtered_options, default=selected_preferences,disabled=st.session_state['editable'] != "Editable",placeholder="")
    st.caption("Count ("+str(len(selected_likes))+")")
    if st.session_state['editable'] == "Editable":
        st.button('💾 Save preferences',on_click=savePreference,args=[selected_likes])

if st.session_state['editable'] == "Editable":
    pass#st.button("Unlock and stop editing",on_click=removeEdit)
elif st.session_state['editable'] == "Not Lock":
    st.button("✏️Edit - Remember to Unlock after editing",on_click=lockAndEdit)
else: #Lock by someone else
    st.write(st.session_state['editable'])

def find_matches(preferences):
    matches = []
    single_likes = []
    checked_pairs = set()
    for person1, liked_people in preferences.items():
        for person2 in liked_people:
            if person2 in preferences and person1 in preferences[person2]:
                pair = tuple(sorted([person1, person2]))
                if pair not in checked_pairs:
                    matches.append(pair)
                    checked_pairs.add(pair)
            else:
                single_likes.append((person1, person2))
    return matches, single_likes
matches, single_likes = find_matches(preferences)


st.write("Perfect matches for "+selected_person+" ("+str(len([match for match in matches if match[0] == selected_person or match[1] == selected_person]))+"):")
st.caption(" ".join(["("+match[0]+"↔"+match[1]+")" for match in matches if match[0] == selected_person or match[1] == selected_person]))
st.write("")
st.markdown("Single-sided likes for "+selected_person+" ("+str(len([single_like for single_like in single_likes if single_like[0] == selected_person or single_like[1] == selected_person]))+"):")
st.caption(" ".join(["("+single_like[0]+"→"+single_like[1]+")" for single_like in single_likes if single_like[0] == selected_person or single_like[1] == selected_person]))



st.divider()
st.header("Generate Schedule")

# Load the Excel file
numOfRoom = 6#st.slider("Number of Room",1,7,6)

st.write("All Perfect matches ("+str(len(matches))+"):")
st.caption(" ".join(["("+match[0]+"↔"+match[1]+")" for match in matches]))
st.write("")
st.write("All Single-sided likes ("+str(len(single_likes))+"):")
st.caption(" ".join(["("+single_like[0]+"→"+single_like[1]+")" for single_like in single_likes]))



if excel_file := st.file_uploader("Upload your Schedule Excel"):
    wb = load_workbook(filename=excel_file, data_only=True)
    sheet = wb[wb.sheetnames[0]]


    # Create a 2D array to store the data
    data = []

    # Iterate through the cells and check if the text is unbold
    unbold_cells = []
    for row in sheet.iter_rows():
        if row[0].value is not None:
            data.append([x.value for x in row])
        for cell in row:
            if (cell.font.bold and cell.value is not None) == False:
                unbold_cells.append(cell)

    #Remove Unbold data
    for unbold in unbold_cells:
        data[unbold.row-1][unbold.column-1] = None
    
    #Handle Merged Cell
    # Get a list of merged cell ranges
    merged_cells_ranges = sheet.merged_cells.ranges

    # Initialize an empty list to store individual merged cells
    merged_cells = []

    # Iterate over each merged cell range
    for merged_range in merged_cells_ranges:
        min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                merged_cells.append(cell)

    # Set content of  merged cells to MERGED and so that later will not overwrite if content is MERGED
    for cell in merged_cells:
        data[cell.row-1][cell.column-1] = "MeRGeD"


    output_dict = {element: index for index, element in enumerate(data[0][1:], start=1)}
    reversed_dict = {v: k for k, v in output_dict.items()}
    print(reversed_dict)


    def SetMeetingFor(column1,column2,data,error):

   
        # Find the rows where both given columns are None
        index = None
        roomNum = 0
        for i, row in enumerate(data):
            if row[column1] is None and row[column2] is None:
                #if sum(1 for s in row if "↔" in str(s) and "*" not in str(s)) < (numOfRoom * 2):
                if column1 == output_dict['Alpha'] or column1 == output_dict['Bravo']:
                    if (row[output_dict['Room 0']] is None):
                        index = i
                        roomNum = "Room 0"
                        break
                else:
                    if (row[output_dict['Room 1']] is None):
                        index = i
                        roomNum = "Room 1"
                        break
                    if (row[output_dict['Room 2']] is None):
                        index = i
                        roomNum = "Room 2"
                        break
                    if (row[output_dict['Room 3']] is None):
                        index = i
                        roomNum = "Room 3"
                        break
                    if (row[output_dict['Room 4']] is None):
                        index = i
                        roomNum = "Room 4"
                        break
                    if (row[output_dict['Room 5']] is None):
                        index = i
                        roomNum = "Room 5"
                        break
                    if (row[output_dict['Room 6']] is None):
                        index = i
                        roomNum = "Room 6"
                        break

        # Print the result

        if (index):
            data[index][column1] = reversed_dict[column1] + "↔"+ reversed_dict[column2] + "\n"+ roomNum
            data[index][column2] = reversed_dict[column1] + "↔"+ reversed_dict[column2] + "\n"+ roomNum
            data[index][output_dict[roomNum]] = reversed_dict[column1] + "↔"+ reversed_dict[column2] + "\n"+ roomNum
            # for index, row in rows_with_none:
            #     print(f"Row {index + 1}: {row}")
        else:
            error += ["Error not match slot / room for "+ reversed_dict[column1] + "↔"+ reversed_dict[column2]]

        return  data,error





    def permutations_fun(arr):
        permutations = [random.sample(arr, len(arr)) for _ in range(num_permutations)]
        return permutations
    #Remove matches that already got bold
    flattened_array = [item for sublist in data for item in sublist]
    flattened_array_split = [s.split("\n")[0] if "\n" in str(s) else s for s in flattened_array]
    for match in matches:
        # Count the occurrences of "ABC"
        count = flattened_array_split.count(match[0] + "↔"+ match[1]+"*") + flattened_array_split.count(match[1] + "↔"+ match[0]+"*")
        if count == 2:
            print("Remove ("+match[0] + "↔"+ match[1]+")")
            matches.remove(match)
        count = flattened_array_split.count(match[0] + "↔"+ match[1])+flattened_array_split.count(match[1] + "↔"+ match[0])
        if count == 3:
            print("Remove ("+match[0] + "↔"+ match[1]+")")
            matches.remove(match)

    with st.spinner("Trying many combination on your behalf"):
        perms = permutations_fun(matches)
        perm_list = list(perms)
        best_perm = None
        best_error = float('inf')  # Set initial best error to infinity
        for perm in perm_list:
            errors = []
            tempdata = copy.deepcopy(data)
            for match in perm:
                tempdata,errors = SetMeetingFor(output_dict[match[0]], output_dict[match[1]], tempdata, errors)
            if len(errors) < best_error:  # Check if the current error is lower than the best error
                best_perm = perm  # Update best_perm
                best_error = len(errors)  # Update best_error
            if best_error == 0:  # Break if error is 0
                break



    from copy import copy

    print("Best permutation:", best_perm)
    print("Lowest error:", best_error)


    errors = []
    for match in best_perm:
        data,errors = SetMeetingFor(output_dict[match[0]], output_dict[match[1]], data, errors)

    for error in errors:
        st.error(error)
    st.caption("Error Count ("+str(len(errors))+")")


    # Iterate through the 2d Array and write to the Excel file
    for r, row in enumerate(data, start=1):
        for c, value in enumerate(row, start=1):
            try:
                if value != "MeRGeD":
                    sheet.cell(row=r, column=c).value = value
            except:
                pass
            
    for cell in unbold_cells:
        sheet[cell.coordinate].font = Font(bold=False)
        sheet[cell.coordinate].alignment = Alignment(wrapText=True, vertical='center',horizontal="center")
    # Save the Excel file
    wb.save('output.xlsx')



    with open('output.xlsx', "rb") as file:
        st.download_button(
            label="Download",
            data=file,
            file_name='output.xlsx',
            mime='application/msexcel',
        )

st.divider()
st.write("Features")
st.caption("""
1. UI for user to key in Preference
2. UI show for each person, who they single sided likes
3. UI show for each person, who they matched with
4. Read and understand schedule from excel
5. Bold mean confirmed event, it will not touch
6. Unbold event mean event that can shift
7. Ensure only X meeting happening at one time based on Number of Room
8. Use * to denote that the meeting is happen external and it will not take up room
9. Only Meeting with Alpha and Bravo can use meeting Room 0
""")
st.write("")
st.write("Algorithm in Text")
st.caption("""
1. Using data collected from the perferences
2. The algorithm will find a list of two way matches
3. After you upload the excel schedule      
4. It will convert the spreadsheet into a 2d array      
5. All unbold cell will be cleared 
6. It will treat merged cell as used timeslot (Same as Bolded)
7. For all the matched pairing, if there is already a bolded pairing, it will not be added
8. However, it will count as thou it will take up a room space    
9. If the cell is ends with an *, it is consider that the meeting place is external
10. The system will randomly try 50000 permutation to try if it can fit everyone into the schedule, if it fail, the least error will be used    
11. It will keep the integrity of ensuring room limit not exceeded, and it will not take a timeslot that is bold or merged.""")
